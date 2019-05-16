
from openpyxl import load_workbook
from openpyxl import  workbook


workbook = load_workbook('./res/各大机构2019暑秋价格.xlsx')

print(workbook.sheetnames)

sheet = workbook[workbook.sheetnames[0]];

print(sheet.title);

for row in range(2,5):
    for col in range(4, 5):
        cell_index = chr(col) + str(row);
        print(sheet[cell_index].value, end='\t')

    print()







